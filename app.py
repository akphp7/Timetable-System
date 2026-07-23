# ============================================================================
# app.py — The main Flask application for the College Timetable Generator
# ============================================================================
#
# WHAT THIS FILE DOES:
# This is the heart of the entire project. It:
#   1. Creates a Flask web server
#   2. Defines the database tables (models)
#   3. Contains helper functions for scheduling logic
#   4. Defines all the URL routes (pages) the user can visit
#
# HOW A FLASK APP WORKS:
#   - Flask is a Python web framework. You define "routes" (URLs), and when
#     a user visits that URL, Flask calls the corresponding Python function.
#   - That function returns HTML (usually via a template) which the browser
#     renders as a web page.
# ============================================================================

# ---------------------------------------------------------------------------
# SECTION 1: IMPORTS
# ---------------------------------------------------------------------------
# Each import brings in a library/module that gives us specific capabilities.

from flask import (
    Flask,              # The core framework — creates the web application
    render_template,    # Turns HTML template files into actual HTML pages
    request,            # Gives access to data sent by the user (form data, URL params)
    redirect,           # Sends the user to a different page
    url_for,            # Generates URLs for our routes by name (avoids hardcoding)
    flash,              # Shows one-time messages to the user (success/error alerts)
    send_file,          # Sends a file (like Excel/ZIP) as a download to the user
    session             # Stores data per-user (like login state) using browser cookies
)
from flask_sqlalchemy import SQLAlchemy  # ORM — lets us use Python classes instead of raw SQL
from functools import wraps              # Used to create decorators (like @login_required)
import random                            # For randomly picking available time slots
import os                                # For file/directory operations and environment variables
import csv                               # For reading CSV files (classroom/lab bulk upload)
from io import TextIOWrapper, BytesIO    # For handling file streams in memory
import openpyxl                          # For creating Excel (.xlsx) files
from openpyxl.styles import Font         # For styling Excel cells (bold, size, etc.)
from openpyxl.utils import get_column_letter  # Converts column number to letter (1→A, 2→B)
from pathlib import Path                 # Modern way to handle file paths
from zipfile import ZipFile              # For creating ZIP archives (multiple Excel downloads)
from werkzeug.security import (
    generate_password_hash,   # Hashes passwords before storing (security!)
    check_password_hash       # Verifies a password against its hash
)
from urllib.parse import urlparse, urljoin  # For validating redirect URLs (security!)

# ---------------------------------------------------------------------------
# SECTION 2: APP CONFIGURATION
# ---------------------------------------------------------------------------
# Here we create the Flask app and configure it.

app = Flask(__name__)
# __name__ tells Flask where to find templates, static files, etc.

# Configuration settings:
app.config['UPLOAD_FOLDER'] = 'uploads'
# Where uploaded CSV files would be stored (not heavily used in this app)

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///timetable.db'
# This tells SQLAlchemy to use SQLite (a simple file-based database).
# The file will be created at instance/timetable.db automatically.
# SQLite is perfect for small apps — no separate database server needed.

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# Disables a Flask-SQLAlchemy feature that tracks every change to objects.
# We don't need it, and it wastes memory.

app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
# SECRET_KEY is used to sign session cookies and flash messages.
# In production, you should set a real random key via environment variable.
# The fallback 'dev-secret-key-change-in-production' is only for development.

db = SQLAlchemy(app)
# Creates the SQLAlchemy database object, linked to our Flask app.
# From now on, we use `db` to define tables and run queries.


# ============================================================================
# SECTION 3: DATABASE MODELS
# ============================================================================
# Models are Python classes that represent database tables.
# Each class = one table. Each class attribute = one column.
# SQLAlchemy automatically creates the SQL tables from these classes.
#
# Think of it like this:
#   Class Classroom → creates a table called "classroom"
#   id, name, capacity → creates columns in that table

class Classroom(db.Model):
    """
    Represents a physical classroom where lectures can be held.
    Example row: id=1, name="LH-101", capacity=60
    """
    id = db.Column(db.Integer, primary_key=True)        # Auto-incrementing unique ID
    name = db.Column(db.String(50), unique=True, nullable=False)  # Room name, must be unique
    capacity = db.Column(db.Integer, nullable=False)     # How many students it can hold


class Lab(db.Model):
    """
    Represents a computer/science lab where practical sessions are held.
    Labs are separate from classrooms because lab sessions are 2 hours.
    """
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    capacity = db.Column(db.Integer, nullable=False)


class Batch(db.Model):
    """
    A batch is a group of students (e.g., "B.Tech 2024 CSE").
    Each batch gets its own timetable.
    """
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)


class Course(db.Model):
    """
    A subject/course that needs to be scheduled.
    
    Fields:
    - credits: How many hours per week this course needs (e.g., 3 = three 1-hour slots)
    - is_lab: If True, this is a lab course (gets 2 consecutive hours)
    - priority: If True, this course uses a special scheduling pattern
    """
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    credits = db.Column(db.Integer, nullable=False)
    is_lab = db.Column(db.Boolean, default=False)
    priority = db.Column(db.Boolean, default=False)


class Professor(db.Model):
    """
    A teacher who teaches one or more courses.
    We track professors to avoid double-booking them.
    """
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)


class Schedule(db.Model):
    """
    THE CORE TABLE — each row represents one time slot in one batch's timetable.
    
    For example: "Batch CSE-2024 has Data Structures with Dr. Smith on Monday
                  at slot 2 (10-11 AM) in classroom LH-101"
    
    Fields:
    - batch_id: Which batch this entry belongs to
    - course_id: Which course is being taught
    - professor_id: Which professor is teaching
    - day: 0=Monday, 1=Tuesday, 2=Wednesday, 3=Thursday, 4=Friday
    - slot: 0-8 representing 9 one-hour slots from 8AM to 5PM
    - classroom_id: Which room (NULL if it's a lab session)
    - lab_id: Which lab (NULL if it's a regular lecture)
    
    The foreign keys (db.ForeignKey) link this table to the other tables,
    creating relationships. For example, batch_id references the Batch table.
    """
    id = db.Column(db.Integer, primary_key=True)
    batch_id = db.Column(db.Integer, db.ForeignKey('batch.id'), nullable=False)
    course_id = db.Column(db.Integer, db.ForeignKey('course.id'), nullable=False)
    professor_id = db.Column(db.Integer, db.ForeignKey('professor.id'), nullable=False)
    day = db.Column(db.Integer, nullable=False)       # 0-4 for Monday-Friday
    slot = db.Column(db.Integer, nullable=False)       # 0-8 for time slots
    classroom_id = db.Column(db.Integer, db.ForeignKey('classroom.id'), nullable=True)
    lab_id = db.Column(db.Integer, db.ForeignKey('lab.id'), nullable=True)


class User(db.Model):
    """
    Stores login credentials with role-based access control.
    
    Roles:
      - 'admin'   : Full access — manage batches, classrooms, labs, schedules
      - 'faculty'  : View timetables, download, analytics (read-only)
      - 'student'  : View timetables, download, analytics (read-only)
    
    SECURITY NOTE: We never store passwords in plain text!
    Instead, we store a "hash" — a one-way mathematical transformation.
    """
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='admin')

    def set_password(self, password):
        """Hash and store a password. Called when creating a user."""
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        """Verify a password against the stored hash. Called during login."""
        return check_password_hash(self.password_hash, password)


# ============================================================================
# SECTION 4: AUTHENTICATION HELPERS
# ============================================================================

def login_required(f):
    """
    A DECORATOR that protects routes so only logged-in users can access them.
    
    HOW IT WORKS:
    1. When you write @login_required above a route function, this decorator
       wraps that function.
    2. Before the actual route code runs, it checks if 'user_id' is in the
       session (meaning the user is logged in).
    3. If not logged in → redirects to the login page.
    4. If logged in → lets the original function run normally.
    
    Example usage:
        @app.route('/secret-page')
        @login_required
        def secret_page():
            return "You can only see this if logged in!"
    """
    @wraps(f)  # Preserves the original function's name and docstring
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login', next=request.url))
            # `next=request.url` remembers where the user was trying to go,
            # so after login we can send them there.
        return f(*args, **kwargs)
    return decorated_function


def role_required(*roles):
    """
    Decorator that restricts a route to specific user roles.
    Usage: @role_required('admin') or @role_required('admin', 'faculty')
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Please log in to access this page.', 'warning')
                return redirect(url_for('login', next=request.url))
            if session.get('role') not in roles:
                flash('You do not have permission to access this page.', 'error')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def is_safe_redirect_url(target):
    """
    SECURITY: Validates redirect URLs to prevent "open redirect" attacks.
    
    After login, we redirect to a `next` URL. An attacker could craft a link
    like /login?next=https://evil-site.com to trick users.
    
    This function ensures the redirect URL points to our own server.
    """
    ref_url = urlparse(request.host_url)
    test_url = urlparse(urljoin(request.host_url, target))
    return test_url.scheme in ('http', 'https') and ref_url.netloc == test_url.netloc


# ============================================================================
# SECTION 5: SCHEDULING HELPER FUNCTIONS
# ============================================================================
# These functions contain the core logic for the timetable generation algorithm.

def find_available_classroom(day, slot):
    """
    Finds a classroom that is NOT already booked at the given day+slot.
    
    HOW IT WORKS:
    1. Gets all classrooms from the database
    2. For each classroom, checks if there's already a schedule entry
       using that classroom at the same day+slot
    3. Returns the first free classroom, or None if all are booked
    
    Parameters:
        day (int): 0-4 (Monday-Friday)
        slot (int): 0-8 (time slot)
    
    Returns:
        Classroom object or None
    """
    all_classrooms = Classroom.query.all()
    for classroom in all_classrooms:
        existing_schedule = Schedule.query.filter_by(
            classroom_id=classroom.id, day=day, slot=slot
        ).first()
        if not existing_schedule:
            return classroom
    return None


def find_available_lab(day, slot):
    """
    Finds a lab that is free for TWO consecutive slots (since labs are 2 hours).
    
    HOW IT WORKS:
    1. Gets all labs from the database
    2. For each lab, checks BOTH the current slot AND the next slot
    3. If both are free → this lab can be used
    4. Returns the first available lab, or None
    
    Why 2 slots? Lab sessions always occupy 2 consecutive time slots.
    """
    all_labs = Lab.query.all()
    for lab in all_labs:
        current_schedule = Schedule.query.filter_by(
            lab_id=lab.id, day=day, slot=slot
        ).first()
        next_slot = Schedule.query.filter_by(
            lab_id=lab.id, day=day, slot=slot + 1
        ).first()
        if not current_schedule and not next_slot:
            return lab
    return None


def is_slot_available(batch_id, professor_id, day, slot, is_lab=False):
    """
    Checks if a particular time slot is available for scheduling.
    
    A slot is UNAVAILABLE if:
    - The batch already has something scheduled at that day+slot
    - The professor is already teaching another batch at that day+slot
    
    For lab courses (is_lab=True):
    - We check TWO consecutive slots (current and next)
    - The slot must be ≤ 7 (so slot+1 doesn't exceed 8)
    
    Parameters:
        batch_id: Which batch we're scheduling for
        professor_id: Which professor is teaching
        day: 0-4 (Monday-Friday)
        slot: 0-8 (time slot)
        is_lab: True if this is a lab course needing 2 hours
    
    Returns:
        True if the slot is free, False otherwise
    """
    if is_lab:
        if slot > 7:  # Can't start a 2-hour lab at slot 8 (last slot)
            return False
        for offset in range(2):  # Check both slot and slot+1
            existing_schedule = Schedule.query.filter_by(
                batch_id=batch_id, day=day, slot=slot + offset
            ).first()
            professor_schedule = Schedule.query.filter_by(
                professor_id=professor_id, day=day, slot=slot + offset
            ).first()
            if existing_schedule or professor_schedule:
                return False
        return True
    else:
        # Regular course: just check the one slot
        existing_schedule = Schedule.query.filter_by(
            batch_id=batch_id, day=day, slot=slot
        ).first()
        professor_schedule = Schedule.query.filter_by(
            professor_id=professor_id, day=day, slot=slot
        ).first()
        return not (existing_schedule or professor_schedule)


# ============================================================================
# SECTION 6: EXCEL EXPORT FUNCTION
# ============================================================================

def generate_excel(batch_ids):
    """
    Creates an Excel file containing the timetable for a batch.
    
    HOW IT WORKS:
    1. Gets the batch and all its scheduled entries from the database
    2. Builds a 5×9 grid (5 days × 9 time slots) and fills it with course info
    3. Uses the openpyxl library to create a formatted Excel workbook
    4. Saves it to the "timetables" folder and returns the file path
    
    The Excel file looks like a standard timetable grid:
    - Rows = Days (Monday-Friday)
    - Columns = Time slots (8AM-5PM)
    - Cells = Course name, professor, room/lab
    """
    if not batch_ids:
        return None

    batch = Batch.query.get(batch_ids[0])
    if not batch:
        return None

    # Define the time slot labels for Excel column headers
    time_slots = [
        "08:00 AM - 09:00 AM", "09:00 AM - 10:00 AM", "10:00 AM - 11:00 AM",
        "11:00 AM - 12:00 PM", "12:00 PM - 01:00 PM", "01:00 PM - 02:00 PM",
        "02:00 PM - 03:00 PM", "03:00 PM - 04:00 PM", "04:00 PM - 05:00 PM"
    ]
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    # Create a 5×9 grid initialized with empty strings
    timetable = [["" for _ in range(9)] for _ in range(5)]

    # Slot 4 (12:00-1:00 PM) is always lunch
    for day in range(5):
        timetable[day][4] = "Lunch"

    # Fill the grid from database records
    schedules = Schedule.query.filter_by(batch_id=batch.id).all()
    for schedule in schedules:
        if schedule.slot == 4:  # Don't overwrite lunch
            continue
        course = Course.query.get(schedule.course_id)
        professor = Professor.query.get(schedule.professor_id)
        classroom = Classroom.query.get(schedule.classroom_id)
        lab = Lab.query.get(schedule.lab_id)

        # Build display string: "CourseName (ProfName) [LabName] {RoomName}"
        entry = f"{course.name} ({professor.name})"
        if lab:
            entry += f" [{lab.name}]"
        if classroom:
            entry += f" {{{classroom.name}}}"

        timetable[schedule.day][schedule.slot] = entry

    # --- Create the Excel file using openpyxl ---
    wb = openpyxl.Workbook()       # Create a new workbook
    ws = wb.active                 # Get the active (first) worksheet
    ws.title = "TimeTable"

    # Row 1: Title
    ws['A1'] = f"Timetable for Batch: {batch.name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:J1')  # Merge across all columns for a wide title

    # Row 2: Column headers (Day + time slots)
    ws['A2'] = "Day"
    for col_index, time_slot in enumerate(time_slots, start=2):
        col_letter = get_column_letter(col_index)
        ws[f'{col_letter}2'] = time_slot

    # Rows 3-7: The actual timetable data
    for day_index, day_name in enumerate(days):
        ws[f"A{day_index + 3}"] = day_name
        for slot_index, slot_content in enumerate(timetable[day_index]):
            col_letter = get_column_letter(slot_index + 2)
            ws[f"{col_letter}{day_index + 3}"] = slot_content

    # Save to disk
    current_dir = Path(__file__).parent
    timetables_dir = current_dir / "timetables"
    timetables_dir.mkdir(exist_ok=True)
    excel_path = timetables_dir / f"batch_{batch.id}.xlsx"
    wb.save(excel_path)

    return excel_path


# ============================================================================
# SECTION 7: ROUTES — The pages users can visit
# ============================================================================
# Each @app.route(...) maps a URL to a Python function.
# The function processes the request and returns a response (usually HTML).

# ---------- LOGIN / LOGOUT ----------

@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    The login page.
    
    GET request: Shows the login form
    POST request: Processes the form submission (checks username + password)
    
    FLOW:
    1. If user is already logged in → go to home page
    2. If form submitted → look up the user, verify password
    3. If credentials match → store user info in session, redirect to home
    4. If wrong → show error message
    """
    if 'user_id' in session:
        return redirect(url_for('index'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            # Login successful — store user ID, username, and role in session
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            flash(f'Logged in as {user.role}.', 'success')

            # Redirect to the page they originally tried to visit (if any)
            next_url = request.args.get('next')
            if next_url and is_safe_redirect_url(next_url):
                return redirect(next_url)
            return redirect(url_for('index'))

        flash('Invalid username or password.', 'error')

    return render_template('login.html')


@app.route('/logout')
def logout():
    """
    Logs the user out by clearing all session data.
    Then redirects to the login page.
    """
    session.clear()
    flash('Logged out successfully.', 'success')
    return redirect(url_for('login'))


# ---------- HOME PAGE ----------

@app.route('/')
@login_required
def index():
    """
    The home page — shows all batches as cards.
    Each batch has buttons to Manage, Download, or Delete it.
    """
    batches = Batch.query.all()
    return render_template('index.html', batches=batches)


# ---------- BATCH MANAGEMENT ----------

@app.route('/create_batch', methods=['GET', 'POST'])
@role_required('admin')
def create_batch():
    """
    Page to create a new student batch.
    
    GET: Shows a form with a "Batch Name" field
    POST: Creates the batch in the database
    
    Uses try/except because the batch name must be unique —
    if a duplicate is added, the database will throw an error.
    """
    if request.method == 'POST':
        batch_name = request.form['name']
        new_batch = Batch(name=batch_name)
        db.session.add(new_batch)
        try:
            db.session.commit()
            flash('Batch created successfully', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating batch: {str(e)}', 'error')
        return redirect(url_for('index'))

    return render_template('create_batch.html')


@app.route('/batch/<int:batch_id>/delete', methods=['POST'])
@role_required('admin')
def delete_batch(batch_id):
    """
    Deletes a batch AND all its scheduled courses.
    
    <int:batch_id> in the URL means Flask extracts that part of the URL
    and passes it as a parameter. For example, /batch/3/delete → batch_id=3
    
    We first delete all Schedule entries for this batch (otherwise we'd have
    orphan records pointing to a deleted batch), then delete the batch itself.
    """
    batch = Batch.query.get_or_404(batch_id)
    try:
        Schedule.query.filter_by(batch_id=batch_id).delete()
        db.session.delete(batch)
        db.session.commit()
        flash(f'Batch "{batch.name}" deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting batch: {str(e)}', 'error')
    return redirect(url_for('index'))


# ---------- CLASSROOM MANAGEMENT ----------

@app.route('/classrooms', methods=['GET', 'POST'])
@role_required('admin')
def manage_classrooms():
    """
    Page to add/view/delete classrooms. Supports two ways to add:
    
    1. MANUAL: Fill in room name + capacity in the form
    2. CSV UPLOAD: Upload a CSV file with columns: name, capacity
    
    The CSV upload is detected by checking if 'file' is in request.files.
    """
    if request.method == "POST":
        if 'file' in request.files:
            # --- CSV UPLOAD PATH ---
            file = request.files["file"]
            if file.filename != '':
                try:
                    # Wrap the binary file stream as text so csv.reader can read it
                    stream = TextIOWrapper(file.stream)
                    csv_input = csv.reader(stream)
                    next(csv_input)  # Skip the header row

                    success_count = 0
                    duplicate_count = 0
                    invalid_count = 0

                    for row in csv_input:
                        # Validate: need at least 2 non-empty columns
                        if len(row) < 2 or not row[0].strip() or not row[1].strip():
                            invalid_count += 1
                            continue

                        name = row[0].strip()
                        try:
                            capacity = int(row[1])
                        except ValueError:
                            invalid_count += 1
                            continue

                        # Skip duplicates
                        if Classroom.query.filter_by(name=name).first():
                            duplicate_count += 1
                            continue

                        new_classroom = Classroom(name=name, capacity=capacity)
                        db.session.add(new_classroom)
                        success_count += 1

                    db.session.commit()
                    flash_message = f"Successfully added {success_count} classrooms"
                    if duplicate_count:
                        flash_message += f", skipped {duplicate_count} duplicates"
                    if invalid_count:
                        flash_message += f", ignored {invalid_count} invalid rows"
                    flash(flash_message, 'success' if success_count else 'warning')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error processing file: {str(e)}', 'error')
        else:
            # --- MANUAL ADD PATH ---
            name = request.form.get("name")
            capacity = request.form.get("capacity")
            if name and capacity:
                new_classroom = Classroom(name=name, capacity=capacity)
                db.session.add(new_classroom)
                try:
                    db.session.commit()
                    flash('Classroom added successfully', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error adding classroom: {str(e)}', 'error')

        return redirect(url_for("manage_classrooms"))

    classrooms = Classroom.query.all()
    return render_template('classrooms.html', classrooms=classrooms)


@app.route('/classrooms/<int:classroom_id>/delete', methods=['POST'])
@role_required('admin')
def delete_classroom(classroom_id):
    """
    Deletes a classroom.
    
    Before deleting, we set classroom_id=NULL on any schedule entries
    that reference this classroom. This prevents database errors from
    foreign key violations (a schedule can't point to a deleted classroom).
    """
    classroom = Classroom.query.get_or_404(classroom_id)
    try:
        Schedule.query.filter_by(classroom_id=classroom_id).update({'classroom_id': None})
        db.session.delete(classroom)
        db.session.commit()
        flash(f'Classroom "{classroom.name}" deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting classroom: {str(e)}', 'error')
    return redirect(url_for('manage_classrooms'))


# ---------- LAB MANAGEMENT ----------

@app.route('/labs', methods=['GET', 'POST'])
@role_required('admin')
def manage_labs():
    """
    Page to add/view/delete labs. Same pattern as classrooms:
    supports manual add and CSV upload.
    """
    if request.method == 'POST':
        if 'file' in request.files:
            file = request.files['file']
            if file.filename != '':
                try:
                    stream = TextIOWrapper(file.stream)
                    csv_input = csv.reader(stream)
                    next(csv_input)

                    success_count = 0
                    duplicate_count = 0
                    invalid_count = 0

                    for row in csv_input:
                        if len(row) < 2 or not row[0].strip() or not row[1].strip():
                            invalid_count += 1
                            continue

                        name = row[0].strip()
                        try:
                            capacity = int(row[1])
                            if capacity <= 0:
                                invalid_count += 1
                                continue

                            if Lab.query.filter_by(name=name).first():
                                duplicate_count += 1
                                continue

                            new_lab = Lab(name=name, capacity=capacity)
                            db.session.add(new_lab)
                            success_count += 1
                        except ValueError:
                            invalid_count += 1
                            continue

                    db.session.commit()
                    flash(
                        f'Successfully added {success_count} labs. '
                        f'{duplicate_count} duplicates skipped. '
                        f'{invalid_count} invalid rows ignored.',
                        'success' if success_count else 'warning'
                    )
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error processing CSV: {str(e)}', 'error')
        else:
            lab_name = request.form.get('lab_name')
            lab_capacity = request.form.get('lab_capacity')

            if lab_name and lab_capacity:
                try:
                    capacity = int(lab_capacity)
                    if capacity <= 0:
                        flash('Capacity must be positive', 'error')
                    elif Lab.query.filter_by(name=lab_name).first():
                        flash('Lab already exists', 'error')
                    else:
                        new_lab = Lab(name=lab_name, capacity=capacity)
                        db.session.add(new_lab)
                        db.session.commit()
                        flash('Lab added successfully', 'success')
                except ValueError:
                    flash('Invalid capacity value', 'error')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error adding lab: {str(e)}', 'error')

        return redirect(url_for('manage_labs'))

    labs = Lab.query.order_by(Lab.name).all()
    return render_template('labs.html', labs=labs)


@app.route('/labs/<int:lab_id>/delete', methods=['POST'])
@role_required('admin')
def delete_lab(lab_id):
    """Deletes a lab. Same pattern as delete_classroom."""
    lab = Lab.query.get_or_404(lab_id)
    try:
        Schedule.query.filter_by(lab_id=lab_id).update({'lab_id': None})
        db.session.delete(lab)
        db.session.commit()
        flash(f'Lab "{lab.name}" deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting lab: {str(e)}', 'error')
    return redirect(url_for('manage_labs'))


# ---------- BATCH TIMETABLE MANAGEMENT (THE BIG ONE!) ----------

@app.route('/batch/<int:batch_id>', methods=['GET', 'POST'])
@login_required
def manage_batch(batch_id):
    # Only admins can POST (schedule new courses)
    if request.method == 'POST' and session.get('role') != 'admin':
        flash('Only admins can schedule courses.', 'error')
        return redirect(url_for('manage_batch', batch_id=batch_id))
    """
    THE MAIN PAGE — View and manage a batch's timetable.
    
    GET: Shows the timetable grid + a form to schedule new courses
    POST: Processes the scheduling form and auto-places the course
    
    =====================================================================
    SCHEDULING ALGORITHM EXPLAINED:
    =====================================================================
    
    When you submit the form, the app needs to decide WHERE to place the
    course in the 5-day × 9-slot grid. The algorithm works like this:
    
    1. COLLECT CONSTRAINTS:
       - Is it a lab? (needs 2 consecutive hours)
       - Is there a priority pattern? (2-hour consecutive, 2-1-1, 2-1)
       - Is there a shift preference? (morning vs afternoon)
       - Is there a day preference? (prefer Monday, etc.)
       - Is there a day to avoid? (skip Wednesday, etc.)
    
    2. FIND AVAILABLE SLOTS:
       - Loop through all days (Monday-Friday)
       - Skip the "avoid day" if set
       - For each day, loop through time slots
       - Skip slot 4 (lunch break)
       - Check if the slot is free (no batch conflict, no professor conflict)
       - Collect all valid (day, slot) pairs
    
    3. PLACE THE COURSE:
       - Based on the scheduling mode, pick slots from the available list
       - Create Schedule database entries
       - Assign a free classroom or lab
    
    =====================================================================
    SCHEDULING MODES:
    =====================================================================
    
    A) REGULAR COURSE (no special options):
       - Randomly pick N slots from available ones (N = credits)
       - Each slot is on a different day
    
    B) LAB COURSE (is_lab checked):
       - Find a day where 2 consecutive free slots exist
       - Place both hours on that day, assign a lab room
    
    C) PRIORITY - "2-hour consecutive":
       - Like a lab but in a regular classroom
       - Places TWO 2-hour blocks on different days
    
    D) PRIORITY - "2-1-1" pattern:
       - 2 consecutive hours on one day
       - 1 hour on a different day
       - 1 hour on yet another day
       - Good for 4-credit courses
    
    E) PRIORITY - "2-1" pattern:
       - 2 consecutive hours on one day
       - 1 hour on a different day
       - Good for 3-credit courses
    
    F) PRIORITY SHIFT:
       - Only schedules in first half (slots 0-3) or second half (slots 5-8)
    
    G) PRIORITY DAY:
       - Tries to place the course on the preferred day first
    """
    batch = Batch.query.get_or_404(batch_id)

    if request.method == 'POST':
        # --- Extract form data ---
        course_name = request.form['course_name']
        credits = int(request.form['credits'])
        is_lab = 'is_lab' in request.form           # Checkbox: present = checked
        priority = 'priority' in request.form
        priority_type = request.form.get('priority_type')
        priority_shift = 'priority_shift' in request.form
        priority_shift_type = request.form.get('priority_shift_type')
        priority_day = 'priority_day' in request.form
        priority_day_type = request.form.get('priority_day_type')
        avoid_day = request.form.get('avoid_day')
        if avoid_day:
            avoid_day = int(avoid_day)

        professor_name = request.form['professor_name']

        # --- Get or create the Course ---
        # If a course with this name already exists, reuse it.
        # Otherwise, create a new one.
        course = Course.query.filter_by(name=course_name).first()
        if not course:
            course = Course(name=course_name, credits=credits, is_lab=is_lab, priority=priority)
            db.session.add(course)

        # --- Get or create the Professor ---
        professor = Professor.query.filter_by(name=professor_name).first()
        if not professor:
            professor = Professor(name=professor_name)
            db.session.add(professor)

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash('Error adding course or professor', 'error')
            return redirect(url_for('manage_batch', batch_id=batch_id))

        # --- Map day names to numbers ---
        day_map = {
            "Monday": 0, "Tuesday": 1, "Wednesday": 2,
            "Thursday": 3, "Friday": 4
        }
        priority_day_index = day_map.get(priority_day_type, -1)

        # ===============================================================
        # STEP 1: FIND AVAILABLE SLOTS based on the scheduling mode
        # ===============================================================
        available_slots = []

        if is_lab:
            # LAB: Need 2 consecutive free slots on the same day
            for day in range(5):
                if day == avoid_day:
                    continue
                for slot in range(8):  # Up to slot 7 (so slot+1 ≤ 8)
                    if is_slot_available(batch.id, professor.id, day, slot, is_lab=True) and slot != 4:
                        available_slots.append((day, slot))
                        break  # One pair per day is enough

        elif priority_day:
            # PRIORITY DAY: Try preferred day first, then others
            for day in range(5):
                if day == avoid_day:
                    continue
                if day == priority_day_type:
                    # For the preferred day, look at all slots
                    for slot in range(9):
                        if is_slot_available(batch.id, professor.id, day, slot, is_lab=True) and slot != 4:
                            available_slots.append((day, slot))
                            break
                else:
                    # For other days, only look at afternoon slots (5-8)
                    for slot in range(5, 9):
                        if is_slot_available(batch.id, professor.id, day, slot, is_lab=True) and slot != 4:
                            available_slots.append((day, slot))
                            break

        elif priority_shift:
            # SHIFT PRIORITY: Only morning (0-4) or afternoon (5-8) slots
            if priority_shift_type == "first_half":
                for day in range(5):
                    if day == avoid_day:
                        continue
                    for slot in range(5):  # Slots 0-4 (but 4 is lunch, skipped)
                        if is_slot_available(batch.id, professor.id, day, slot, is_lab=True) and slot != 4:
                            available_slots.append((day, slot))
                            break
            else:
                for day in range(5):
                    if day == avoid_day:
                        continue
                    for slot in range(5, 9):  # Slots 5-8
                        if is_slot_available(batch.id, professor.id, day, slot, is_lab=True) and slot != 4:
                            available_slots.append((day, slot))
                            break

        elif priority:
            # PRIORITY PATTERNS: Special slot arrangements
            if priority_type == "2-hour consecutive":
                # Find days where 2 consecutive slots are free
                for day in range(5):
                    if day == avoid_day:
                        continue
                    for slot in range(8):
                        if slot != 3 and slot != 4:  # Can't span across lunch
                            if (is_slot_available(batch.id, professor.id, day, slot) and
                                    is_slot_available(batch.id, professor.id, day, slot + 1)):
                                available_slots.append((day, slot))
                                break

            elif priority_type == "2-1-1":
                # First: find a day for the 2-hour block
                for day in range(5):
                    if day == avoid_day:
                        continue
                    for slot in range(8):
                        if slot != 3 and slot != 4:
                            if (is_slot_available(batch.id, professor.id, day, slot) and
                                    is_slot_available(batch.id, professor.id, day, slot + 1) and slot != 4):
                                available_slots.append((day, slot))
                                break
                credits -= 2  # 2 hours accounted for
                # Then: find single slots on different days
                for day in range(5):
                    if day == avoid_day:
                        continue
                    for slot in range(9):
                        if is_slot_available(batch.id, professor.id, day, slot) and slot != 4 and credits > 0:
                            available_slots.append((day, slot))
                            break
                credits -= 1
                for day in range(5):
                    if day == avoid_day:
                        continue
                    for slot in range(9):
                        if is_slot_available(batch.id, professor.id, day, slot) and slot != 4 and credits > 0:
                            available_slots.append((day, slot))
                            break

            elif priority_type == "2-1":
                # First: 2-hour block
                for day in range(5):
                    if day == avoid_day:
                        continue
                    for slot in range(8):
                        if slot != 3 and slot != 4:
                            if (is_slot_available(batch.id, professor.id, day, slot) and
                                    is_slot_available(batch.id, professor.id, day, slot + 1) and slot != 4):
                                available_slots.append((day, slot))
                                break
                credits -= 2
                # Then: 1-hour block on a different day
                for day in range(5):
                    if day == avoid_day:
                        continue
                    for slot in range(9):
                        if is_slot_available(batch.id, professor.id, day, slot) and slot != 4 and credits > 0:
                            available_slots.append((day, slot))
                            break

        else:
            # DEFAULT: Find one free slot per day
            for day in range(5):
                if day == avoid_day:
                    continue
                for slot in range(9):
                    if is_slot_available(batch.id, professor.id, day, slot) and slot != 4:
                        available_slots.append((day, slot))
                        break

        # ===============================================================
        # STEP 2: PLACE THE COURSE into the selected slots
        # ===============================================================

        if priority and priority_type == "2-hour consecutive":
            if not available_slots:
                flash('Not enough consecutive slots available for priority course', 'error')
            else:
                # Pick 2 different days, each getting a 2-hour block
                for _ in range(2):
                    day, start_slot = random.choice(available_slots)
                    classroom = find_available_classroom(day, start_slot)
                    if not classroom:
                        flash('No classrooms available. Please add classrooms first.', 'error')
                        return redirect(url_for('manage_batch', batch_id=batch_id))
                    available_slots.remove((day, start_slot))
                    for offset in range(2):
                        new_schedule = Schedule(
                            batch_id=batch.id, course_id=course.id,
                            professor_id=professor.id, day=day,
                            slot=start_slot + offset, classroom_id=classroom.id
                        )
                        db.session.add(new_schedule)
                    try:
                        db.session.commit()
                        flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
                    except Exception as e:
                        db.session.rollback()
                        flash(f'Error scheduling priority course (2-hour consecutive): {str(e)}', 'error')

        elif priority and priority_type == "2-1-1":
            if len(available_slots) < credits:
                flash('Not enough available slots for 2-1-1 pattern', 'error')
            else:
                # Day 1: 2 consecutive hours
                day1, first_slot = available_slots[0]
                classroom = find_available_classroom(day1, first_slot)
                if not classroom:
                    flash('No classrooms available. Please add classrooms first.', 'error')
                    return redirect(url_for('manage_batch', batch_id=batch_id))
                for offset in range(2):
                    new_schedule = Schedule(
                        batch_id=batch.id, course_id=course.id,
                        professor_id=professor.id, day=day1,
                        slot=first_slot + offset, classroom_id=classroom.id
                    )
                    db.session.add(new_schedule)

                # Days 2 & 3: 1 hour each on different days
                day_count = 0
                for day, slot in available_slots[1:]:
                    classroom = find_available_classroom(day, slot)
                    if not classroom:
                        continue
                    if day != day1 and day_count < 2:
                        new_schedule = Schedule(
                            batch_id=batch.id, course_id=course.id,
                            professor_id=professor.id, day=day,
                            slot=slot, classroom_id=classroom.id
                        )
                        db.session.add(new_schedule)
                        day_count += 1

                try:
                    db.session.commit()
                    flash('Priority course scheduled successfully (2-1-1)', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error scheduling priority course (2-1-1): {str(e)}', 'error')

        elif priority and priority_type == "2-1":
            if len(available_slots) < credits:
                flash('Not enough available slots for 2-1 pattern', 'error')
            else:
                # Day 1: 2 consecutive hours
                day1, first_slot = available_slots[0]
                classroom = find_available_classroom(day1, first_slot)
                if not classroom:
                    flash('No classrooms available. Please add classrooms first.', 'error')
                    return redirect(url_for('manage_batch', batch_id=batch_id))
                for offset in range(2):
                    new_schedule = Schedule(
                        batch_id=batch.id, course_id=course.id,
                        professor_id=professor.id, day=day1,
                        slot=first_slot + offset, classroom_id=classroom.id
                    )
                    db.session.add(new_schedule)

                # Day 2: 1 hour on a different day
                day_count = 0
                for day, slot in available_slots[1:]:
                    classroom = find_available_classroom(day, slot)
                    if not classroom:
                        continue
                    if day != day1 and day_count < 1:
                        new_schedule = Schedule(
                            batch_id=batch.id, course_id=course.id,
                            professor_id=professor.id, day=day,
                            slot=slot, classroom_id=classroom.id
                        )
                        db.session.add(new_schedule)
                        day_count += 1

                try:
                    db.session.commit()
                    flash('Priority course scheduled successfully (2-1)', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error scheduling priority course (2-1): {str(e)}', 'error')

        elif is_lab:
            if not available_slots:
                flash('Not enough available slots for lab', 'error')
            else:
                # Pick a random available day, place 2 consecutive hours
                day, start_slot = random.choice(available_slots)
                lab = find_available_lab(day, start_slot)
                if not lab:
                    flash('No labs available. Please add labs first.', 'error')
                    return redirect(url_for('manage_batch', batch_id=batch_id))
                for offset in range(2):
                    new_schedule = Schedule(
                        batch_id=batch.id, course_id=course.id,
                        professor_id=professor.id, day=day,
                        slot=start_slot + offset, lab_id=lab.id
                    )
                    db.session.add(new_schedule)
                try:
                    db.session.commit()
                    flash('Lab scheduled successfully', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error scheduling lab: {str(e)}', 'error')

        else:
            # DEFAULT: Pick random slots equal to the number of credits
            if len(available_slots) < credits:
                flash('Not enough available slots', 'error')
            else:
                scheduled_slots = random.sample(available_slots, credits)
                for day, slot in scheduled_slots:
                    classroom = find_available_classroom(day, slot)
                    if not classroom:
                        flash('No classrooms available. Please add classrooms first.', 'error')
                        return redirect(url_for('manage_batch', batch_id=batch_id))
                    new_schedule = Schedule(
                        batch_id=batch.id, course_id=course.id,
                        professor_id=professor.id, day=day,
                        slot=slot, classroom_id=classroom.id
                    )
                    db.session.add(new_schedule)
                try:
                    db.session.commit()
                    flash('Course scheduled successfully', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error scheduling course: {str(e)}', 'error')

        return redirect(url_for('manage_batch', batch_id=batch_id))

    # --- GET REQUEST: Build the timetable grid for display ---
    schedules = Schedule.query.filter_by(batch_id=batch_id).all()
    timetable = [["-" for _ in range(9)] for _ in range(5)]

    # Mark lunch break
    for day in range(5):
        timetable[day][4] = "Lunch"

    # Fill in scheduled courses
    for schedule in schedules:
        if schedule.slot == 4:
            continue
        course = Course.query.get(schedule.course_id)
        professor = Professor.query.get(schedule.professor_id)
        classroom = Classroom.query.get(schedule.classroom_id)
        lab = Lab.query.get(schedule.lab_id)

        entry = f"{course.name}({professor.name})"
        if lab is not None:
            entry += f", {lab.name}"
        if classroom is not None:
            entry += f" {classroom.name}"

        timetable[schedule.day][schedule.slot] = entry

    return render_template('manage_batch.html', batch=batch, timetable=timetable, schedules=schedules)


# ---------- TIMETABLE OPERATIONS ----------

@app.route('/batch/<int:batch_id>/clear', methods=['POST'])
@role_required('admin')
def clear_batch_timetable(batch_id):
    """
    Clears ALL scheduled courses for a batch (wipes the timetable clean).
    The batch itself is NOT deleted, just its schedule entries.
    """
    batch = Batch.query.get_or_404(batch_id)
    try:
        Schedule.query.filter_by(batch_id=batch_id).delete()
        db.session.commit()
        flash(f'Timetable for batch "{batch.name}" cleared successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error clearing timetable: {str(e)}', 'error')
    return redirect(url_for('manage_batch', batch_id=batch_id))


@app.route('/schedule/<int:schedule_id>/delete', methods=['POST'])
@role_required('admin')
def delete_schedule(schedule_id):
    """Deletes a single schedule entry (one slot from the timetable)."""
    schedule = Schedule.query.get_or_404(schedule_id)
    batch_id = schedule.batch_id
    try:
        db.session.delete(schedule)
        db.session.commit()
        flash('Schedule entry removed successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error removing schedule entry: {str(e)}', 'error')
    return redirect(url_for('manage_batch', batch_id=batch_id))


@app.route('/schedule/<int:schedule_id>/edit', methods=['GET', 'POST'])
@role_required('admin')
def edit_schedule(schedule_id):
    """
    Allows moving a schedule entry to a different day/slot.
    
    GET: Shows a form with dropdowns for new day and slot
    POST: Validates the move (checks for conflicts) and updates the record
    
    Conflict checking ensures:
    1. The batch doesn't already have something at the new day+slot
    2. The professor isn't teaching elsewhere at the new day+slot
    """
    schedule = Schedule.query.get_or_404(schedule_id)
    batch_id = schedule.batch_id

    if request.method == 'POST':
        new_day = int(request.form['day'])
        new_slot = int(request.form['slot'])

        # Can't move to lunch slot
        if new_slot == 4:
            flash('Slot 4 is reserved for lunch break.', 'error')
            return redirect(url_for('edit_schedule', schedule_id=schedule_id))

        # Check for conflicts (exclude the current entry from the check)
        conflict = Schedule.query.filter(
            Schedule.batch_id == batch_id,
            Schedule.day == new_day,
            Schedule.slot == new_slot,
            Schedule.id != schedule_id
        ).first()
        prof_conflict = Schedule.query.filter(
            Schedule.professor_id == schedule.professor_id,
            Schedule.day == new_day,
            Schedule.slot == new_slot,
            Schedule.id != schedule_id
        ).first()

        if conflict or prof_conflict:
            flash('Selected slot is already occupied.', 'error')
            return redirect(url_for('edit_schedule', schedule_id=schedule_id))

        try:
            schedule.day = new_day
            schedule.slot = new_slot
            classroom = find_available_classroom(new_day, new_slot)
            if classroom:
                schedule.classroom_id = classroom.id
            db.session.commit()
            flash('Schedule entry moved successfully.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error moving schedule entry: {str(e)}', 'error')
        return redirect(url_for('manage_batch', batch_id=batch_id))

    # GET: Show the edit form
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    time_slots = [
        "08:00 AM - 09:00 AM", "09:00 AM - 10:00 AM", "10:00 AM - 11:00 AM",
        "11:00 AM - 12:00 PM", "12:00 PM - 01:00 PM", "01:00 PM - 02:00 PM",
        "02:00 PM - 03:00 PM", "03:00 PM - 04:00 PM", "04:00 PM - 05:00 PM"
    ]
    course = Course.query.get(schedule.course_id)
    return render_template('edit_schedule.html', schedule=schedule, days=days,
                           time_slots=time_slots, course=course)


# ---------- DOWNLOAD / EXPORT ----------

@app.route('/select_batches', methods=['GET'])
@login_required
def select_batches():
    """Shows a page where the user can select batches to download as Excel."""
    batches = Batch.query.all()
    return render_template('select_batches.html', batches=batches)


@app.route('/download-timetable', methods=['POST'])
@login_required
def download_timetable():
    """
    Generates Excel files for selected batches and downloads them as a ZIP.
    
    FLOW:
    1. Get the list of selected batch IDs from the form
    2. For each batch, generate an Excel file
    3. Add all Excel files to a ZIP archive (in memory)
    4. Send the ZIP to the user's browser as a download
    5. Clean up the temporary Excel files
    """
    selected_batch_ids = request.form.getlist('batch_ids[]')
    if not selected_batch_ids:
        flash('No batches selected', 'error')
        return redirect(url_for('index'))

    # Create a ZIP file in memory (not on disk)
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, 'w') as zip_file:
        for batch_id in selected_batch_ids:
            excel_path = generate_excel([int(batch_id)])
            if excel_path and os.path.exists(excel_path):
                zip_file.write(excel_path, os.path.basename(excel_path))
                os.remove(excel_path)  # Clean up temp file

    zip_buffer.seek(0)  # Rewind to the start so send_file reads from the beginning
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name='timetables.zip'
    )


# ---------- ANALYTICS DASHBOARD ----------

@app.route('/analytics')
@login_required
def analytics():
    """
    Dashboard showing utilization stats, professor workloads,
    and scheduling metrics across all batches.
    """
    # --- Basic counts ---
    total_batches = Batch.query.count()
    total_classrooms = Classroom.query.count()
    total_labs = Lab.query.count()
    total_courses = Course.query.count()
    total_professors = Professor.query.count()
    total_schedules = Schedule.query.count()

    # --- Classroom utilization ---
    # Total possible slots = classrooms × 5 days × 8 usable slots (excluding lunch)
    total_classroom_slots = total_classrooms * 5 * 8
    used_classroom_slots = Schedule.query.filter(Schedule.classroom_id.isnot(None)).count()
    classroom_utilization = round((used_classroom_slots / total_classroom_slots * 100), 1) if total_classroom_slots else 0

    # --- Lab utilization ---
    total_lab_slots = total_labs * 5 * 8
    used_lab_slots = Schedule.query.filter(Schedule.lab_id.isnot(None)).count()
    lab_utilization = round((used_lab_slots / total_lab_slots * 100), 1) if total_lab_slots else 0

    # --- Per-classroom usage ---
    classroom_usage = []
    for cr in Classroom.query.order_by(Classroom.name).all():
        count = Schedule.query.filter_by(classroom_id=cr.id).count()
        classroom_usage.append({'name': cr.name, 'slots_used': count, 'capacity': cr.capacity})

    # --- Per-lab usage ---
    lab_usage = []
    for lb in Lab.query.order_by(Lab.name).all():
        count = Schedule.query.filter_by(lab_id=lb.id).count()
        lab_usage.append({'name': lb.name, 'slots_used': count, 'capacity': lb.capacity})

    # --- Professor workload (total teaching hours per week) ---
    professor_workload = []
    for prof in Professor.query.order_by(Professor.name).all():
        hours = Schedule.query.filter_by(professor_id=prof.id).count()
        professor_workload.append({'name': prof.name, 'hours': hours})
    professor_workload.sort(key=lambda x: x['hours'], reverse=True)

    # --- Courses per batch ---
    batch_course_counts = []
    for batch in Batch.query.order_by(Batch.name).all():
        # Count distinct courses scheduled for this batch
        course_ids = db.session.query(Schedule.course_id).filter_by(batch_id=batch.id).distinct().count()
        total_slots = Schedule.query.filter_by(batch_id=batch.id).count()
        batch_course_counts.append({
            'name': batch.name,
            'courses': course_ids,
            'total_slots': total_slots
        })

    # --- Day-wise distribution (how many classes on each day across all batches) ---
    day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    day_distribution = []
    for day_num in range(5):
        count = Schedule.query.filter_by(day=day_num).count()
        day_distribution.append({'day': day_names[day_num], 'count': count})

    # --- Slot-wise distribution (which time slots are busiest) ---
    slot_labels = [
        "8-9 AM", "9-10 AM", "10-11 AM", "11-12 PM",
        "12-1 PM", "1-2 PM", "2-3 PM", "3-4 PM", "4-5 PM"
    ]
    slot_distribution = []
    for slot_num in range(9):
        count = Schedule.query.filter_by(slot=slot_num).count()
        slot_distribution.append({'slot': slot_labels[slot_num], 'count': count})

    return render_template('analytics.html',
        total_batches=total_batches,
        total_classrooms=total_classrooms,
        total_labs=total_labs,
        total_courses=total_courses,
        total_professors=total_professors,
        total_schedules=total_schedules,
        classroom_utilization=classroom_utilization,
        lab_utilization=lab_utilization,
        classroom_usage=classroom_usage,
        lab_usage=lab_usage,
        professor_workload=professor_workload,
        batch_course_counts=batch_course_counts,
        day_distribution=day_distribution,
        slot_distribution=slot_distribution
    )


# ============================================================================
# SECTION 8: APPLICATION STARTUP
# ============================================================================

if __name__ == '__main__':
    """
    This block runs when you execute: python app.py
    
    It does three things:
    1. Creates all database tables (if they don't exist)
    2. Creates a default admin user (username: admin, password: admin123)
    3. Starts the Flask development server on http://127.0.0.1:5000
    """
    with app.app_context():
        db.create_all()  # Creates tables from all the Model classes above

        # Create default admin if no users exist
        if not User.query.first():
            admin = User(username='admin', role='admin')
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()

    app.run(debug=True, port=5001)
    # debug=True enables:
    #   - Auto-reload when code changes
    #   - Detailed error pages in the browser
    #   - NEVER use debug=True in production!
